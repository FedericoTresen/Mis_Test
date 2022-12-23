import openpyxl


class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file, path, sheetname):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        return (sheet.max_row)


    def getColumnCount(file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        return (sheet.max_column)


    def readData(file, path, sheetname, rownum, columnum):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        return sheet.cell(row = rownum, column = columnum).value


    def writeData(file, path, sheetname, rownum, columnum, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        sheet.cell(row = rownum, column = columnum).value = data
        workbook.save(path)

